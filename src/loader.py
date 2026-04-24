"""Phase 3 — Load: Write a fully styled, multi-sheet Excel report."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).parent.parent / "output"

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT     = "E8A838"
GREEN      = "1E8449"
RED        = "C0392B"
LIGHT_GRAY = "F2F2F2"
WHITE      = "FFFFFF"
DARK_TEXT  = "1A1A2E"


def _border(style="thin", color="BDBDBD"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, col, value, bg=DARK_BLUE, fg=WHITE, bold=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()
    return cell


def _plain(ws, row, col, value, bold=False, align="left", size=10, color=DARK_TEXT, bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def _currency(ws, row, col, value, bold=False, bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = "#,##0.00"
    cell.font = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = _border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def _pct(ws, row, col, value, color=DARK_TEXT, bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = "0.0%"
    cell.font = Font(name="Arial", size=10, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


# ── Sheet 1: Dashboard ────────────────────────────────────────────────────────
def _write_dashboard(ws, kpis: dict, monthly: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    # Title
    ws.merge_cells("B1:K1")
    c = ws["B1"]
    c.value = "SALES PERFORMANCE DASHBOARD — 2024"
    c.font = Font(name="Arial", bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("B2:K2")
    c2 = ws["B2"]
    c2.value = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}   |   Completed orders only"
    c2.font = Font(name="Arial", size=9, color="888888")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # KPI Cards — 2 rows x 4 cards
    kpi_items = [
        ("Total Revenue",     f"${kpis['Total Revenue']:,.0f}",      DARK_BLUE),
        ("Total Orders",      f"{kpis['Total Orders']:,}",           MID_BLUE),
        ("Total Units",       f"{kpis['Total Units']:,}",            GREEN),
        ("Avg Order Value",   f"${kpis['Avg Order Value']:,.2f}",    ACCENT),
        ("Top Region",        kpis["Top Region"],                    "6C3483"),
        ("Top Category",      kpis["Top Category"],                  "117A65"),
        ("Best Month",        kpis["Best Month"],                    "784212"),
        ("Best Month Rev",    f"${kpis['Best Month Revenue']:,.0f}", "922B21"),
    ]
    col_starts = [2, 5, 8, 11]

    for idx, (label, value, color) in enumerate(kpi_items):
        row_start  = 4 if idx < 4 else 8
        col_start  = col_starts[idx % 4]
        end_col    = col_start + 2

        ws.merge_cells(start_row=row_start, start_column=col_start,
                       end_row=row_start,   end_column=end_col)
        lc = ws.cell(row=row_start, column=col_start, value=label.upper())
        lc.font      = Font(name="Arial", size=8, bold=True, color=WHITE)
        lc.fill      = PatternFill("solid", fgColor=color)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_start].height = 18

        ws.merge_cells(start_row=row_start+1, start_column=col_start,
                       end_row=row_start+2,   end_column=end_col)
        vc = ws.cell(row=row_start+1, column=col_start, value=value)
        vc.font      = Font(name="Arial", size=14, bold=True, color=color)
        vc.fill      = PatternFill("solid", fgColor=LIGHT_GRAY)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_start+1].height = 28

    # Monthly mini-table
    ws.row_dimensions[12].height = 10
    ws.merge_cells("B13:K13")
    th = ws["B13"]
    th.value     = "MONTHLY REVENUE TREND"
    th.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    th.fill      = PatternFill("solid", fgColor=MID_BLUE)
    th.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 22

    headers  = ["Month", "Revenue ($)", "Units", "Orders", "Avg Order ($)", "MoM Growth"]
    col_map  = [2, 3, 5, 6, 8, 10]
    widths   = [14, 14, 10, 10, 14, 12]
    for h, c, w in zip(headers, col_map, widths):
        _hdr(ws, 14, c, h, bg=DARK_BLUE, size=9)
        ws.column_dimensions[get_column_letter(c)].width = w

    data_start = 15
    for i, row in monthly.iterrows():
        r  = data_start + i
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        _plain(ws, r, 2, row["Month_Label"], bg=bg)
        _currency(ws, r, 3, row["Total_Revenue"], bg=bg)
        _plain(ws, r, 5, int(row["Total_Units"]),  align="right", bg=bg)
        _plain(ws, r, 6, int(row["Total_Orders"]), align="right", bg=bg)
        _currency(ws, r, 8, row["Avg_Order_Value"], bg=bg)
        if pd.notna(row["MoM_Growth"]):
            clr = GREEN if row["MoM_Growth"] > 0 else RED
            _pct(ws, r, 10, row["MoM_Growth"], color=clr, bg=bg)
        else:
            _plain(ws, r, 10, "—", align="center", bg=bg)

    total_row = data_start + len(monthly)
    _plain(ws, total_row, 2, "TOTAL", bold=True, bg=LIGHT_BLUE)
    t = ws.cell(total_row, 3, value=f"=SUM(C{data_start}:C{total_row-1})")
    t.number_format = "#,##0.00"
    t.font = Font(name="Arial", bold=True, size=10)
    t.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    t.alignment = Alignment(horizontal="right")
    t.border = _border()


# ── Sheet 2: Monthly Summary ──────────────────────────────────────────────────
def _write_monthly(ws, monthly: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.merge_cells("B1:H1")
    t = ws["B1"]
    t.value     = "MONTHLY SALES SUMMARY"
    t.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    cols   = ["Month", "Total Revenue ($)", "MoM Growth", "Units Sold", "Orders", "Avg Order ($)"]
    widths = [14, 18, 14, 12, 10, 16]
    for i, (h, w) in enumerate(zip(cols, widths), start=2):
        _hdr(ws, 2, i, h, size=10)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 22

    data_row = 3
    for idx, row in monthly.iterrows():
        r  = data_row + idx
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        _plain(ws, r, 2, row["Month_Label"], bg=bg)
        _currency(ws, r, 3, row["Total_Revenue"], bg=bg)
        if pd.notna(row["MoM_Growth"]):
            clr = GREEN if row["MoM_Growth"] > 0 else RED
            _pct(ws, r, 4, row["MoM_Growth"], color=clr, bg=bg)
        else:
            _plain(ws, r, 4, "—", align="center", bg=bg)
        _plain(ws, r, 5, int(row["Total_Units"]),  align="right", bg=bg)
        _plain(ws, r, 6, int(row["Total_Orders"]), align="right", bg=bg)
        _currency(ws, r, 7, row["Avg_Order_Value"], bg=bg)
        ws.row_dimensions[r].height = 18

    last  = data_row + len(monthly) - 1
    total = data_row + len(monthly)
    _plain(ws, total, 2, "TOTAL / AVERAGE", bold=True, bg=LIGHT_BLUE)
    for c in range(3, 8):
        ws.cell(total, c).fill   = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(total, c).border = _border()
    t = ws.cell(total, 3, value=f"=SUM(C{data_row}:C{last})")
    t.number_format = "#,##0.00"
    t.font = Font(name="Arial", bold=True, size=10)
    t.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    t.alignment = Alignment(horizontal="right")
    t.border = _border()
    ws.row_dimensions[total].height = 20

    # Line chart
    chart = LineChart()
    chart.title          = "Monthly Revenue Trend"
    chart.y_axis.title   = "Revenue ($)"
    chart.x_axis.title   = "Month"
    chart.height, chart.width = 12, 22
    chart.style = 2
    rev_data = Reference(ws, min_col=3, min_row=data_row, max_row=last)
    labels   = Reference(ws, min_col=2, min_row=data_row, max_row=last)
    chart.add_data(rev_data)
    chart.set_categories(labels)
    chart.series[0].graphicalProperties.line.solidFill = MID_BLUE
    ws.add_chart(chart, f"B{total + 3}")


# ── Sheet 3: Region Performance ───────────────────────────────────────────────
def _write_regions(ws, region_perf: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.merge_cells("B1:J1")
    t = ws["B1"]
    t.value     = "REGION PERFORMANCE REPORT"
    t.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    cols   = ["Rank", "Region", "Manager", "Target ($)", "Actual Revenue ($)", "Achievement %", "Units", "Orders"]
    widths = [8, 14, 16, 16, 20, 16, 10, 10]
    for i, (h, w) in enumerate(zip(cols, widths), start=2):
        _hdr(ws, 2, i, h, size=10)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 22

    data_row = 3
    for idx, row in region_perf.iterrows():
        r  = data_row + idx
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        _plain(ws, r, 2, int(row["Rank"]),   align="center", bg=bg)
        _plain(ws, r, 3, row["Region Name"], bg=bg)
        _plain(ws, r, 4, row["Manager"],     bg=bg)
        _currency(ws, r, 5, row["Target Revenue"], bg=bg)
        _currency(ws, r, 6, row["Actual_Revenue"],  bg=bg)
        clr = GREEN if row["Target_Achievement"] >= 1 else RED
        _pct(ws, r, 7, row["Target_Achievement"], color=clr, bg=bg)
        _plain(ws, r, 8, int(row["Total_Units"]),  align="right", bg=bg)
        _plain(ws, r, 9, int(row["Total_Orders"]), align="right", bg=bg)
        ws.row_dimensions[r].height = 18

    last  = data_row + len(region_perf) - 1
    total = data_row + len(region_perf)
    _plain(ws, total, 2, "TOTAL", bold=True, bg=LIGHT_BLUE)
    for c in range(3, 10):
        ws.cell(total, c).fill   = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(total, c).border = _border()
    t = ws.cell(total, 6, value=f"=SUM(F{data_row}:F{last})")
    t.number_format = "#,##0.00"
    t.font = Font(name="Arial", bold=True, size=10)
    t.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    t.alignment = Alignment(horizontal="right")
    t.border = _border()
    ws.row_dimensions[total].height = 20

    # Bar chart
    chart = BarChart()
    chart.type           = "col"
    chart.title          = "Revenue by Region"
    chart.y_axis.title   = "Revenue ($)"
    chart.height, chart.width = 12, 20
    chart.style = 2
    data   = Reference(ws, min_col=6, min_row=2, max_row=last)
    labels = Reference(ws, min_col=3, min_row=data_row, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.series[0].graphicalProperties.solidFill = MID_BLUE
    ws.add_chart(chart, f"B{total + 3}")


# ── Sheet 4: Product Breakdown ────────────────────────────────────────────────
def _write_products(ws, product_breakdown: pd.DataFrame, category_summary: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.merge_cells("B1:J1")
    t = ws["B1"]
    t.value     = "PRODUCT & CATEGORY BREAKDOWN"
    t.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Category table
    ws.merge_cells("B2:F2")
    s = ws["B2"]
    s.value     = "BY CATEGORY"
    s.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    s.fill      = PatternFill("solid", fgColor=MID_BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")

    cat_cols = ["Category", "Revenue ($)", "Revenue Share", "Units", "Orders"]
    cat_w    = [18, 16, 15, 12, 10]
    for i, (h, w) in enumerate(zip(cat_cols, cat_w), start=2):
        _hdr(ws, 3, i, h, size=10)
        ws.column_dimensions[get_column_letter(i)].width = w

    cat_start = 4
    for idx, row in category_summary.iterrows():
        r  = cat_start + idx
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        _plain(ws, r, 2, row["Category"], bg=bg)
        _currency(ws, r, 3, row["Revenue"], bg=bg)
        _pct(ws, r, 4, row["Revenue_Share"], bg=bg)
        _plain(ws, r, 5, int(row["Units"]),  align="right", bg=bg)
        _plain(ws, r, 6, int(row["Orders"]), align="right", bg=bg)
        ws.row_dimensions[r].height = 18

    cat_last     = cat_start + len(category_summary) - 1
    prod_hdr_row = cat_last + 3

    # Product table
    ws.merge_cells(f"B{prod_hdr_row}:J{prod_hdr_row}")
    ph = ws[f"B{prod_hdr_row}"]
    ph.value     = "BY PRODUCT"
    ph.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    ph.fill      = PatternFill("solid", fgColor=MID_BLUE)
    ph.alignment = Alignment(horizontal="center", vertical="center")

    prod_cols = ["Category", "Product ID", "Product Name", "Unit Price ($)", "Margin %",
                 "Units Sold", "Revenue ($)", "Revenue Share", "Orders"]
    prod_w    = [14, 12, 22, 14, 12, 12, 16, 15, 10]
    col_hdr   = prod_hdr_row + 1
    for i, (h, w) in enumerate(zip(prod_cols, prod_w), start=2):
        _hdr(ws, col_hdr, i, h, size=9)
        ws.column_dimensions[get_column_letter(i)].width = w

    prod_start = col_hdr + 1
    for idx, row in product_breakdown.reset_index(drop=True).iterrows():
        r  = prod_start + idx
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        _plain(ws, r, 2, row["Category"],    bg=bg)
        _plain(ws, r, 3, row["Product ID"],  align="center", bg=bg)
        _plain(ws, r, 4, row["Product Name"], bg=bg)
        _currency(ws, r, 5, row["Unit Price"], bg=bg)
        _pct(ws, r, 6, row["Margin %"] / 100, bg=bg)
        _plain(ws, r, 7, int(row["Units_Sold"]), align="right", bg=bg)
        _currency(ws, r, 8, row["Revenue"], bg=bg)
        _pct(ws, r, 9, row["Revenue_Share"], bg=bg)
        _plain(ws, r, 10, int(row["Orders"]), align="right", bg=bg)
        ws.row_dimensions[r].height = 18


# ── Sheet 5: Raw Data ─────────────────────────────────────────────────────────
def _write_raw(ws, raw: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.merge_cells("B1:M1")
    t = ws["B1"]
    t.value     = "RAW TRANSACTION DATA  (Completed Orders)"
    t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    display_cols = ["Order ID", "Order Date", "Product Name", "Category",
                    "Region Name", "Sales Rep", "Units", "Unit Price",
                    "Discount %", "Net Revenue", "Payment Method", "Status"]
    col_widths   = [14, 13, 20, 14, 12, 16, 8, 12, 12, 14, 16, 12]

    for i, (h, w) in enumerate(zip(display_cols, col_widths), start=2):
        _hdr(ws, 2, i, h, size=9)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 20

    raw_display = raw[display_cols].reset_index(drop=True)
    data_start  = 3
    for idx, row in raw_display.iterrows():
        r  = data_start + idx
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        for ci, col in enumerate(display_cols, start=2):
            val  = row[col]
            cell = ws.cell(r, ci)
            if col == "Order Date":
                cell.value = val.strftime("%Y-%m-%d") if hasattr(val, "strftime") else str(val)
            elif col in ("Unit Price", "Net Revenue"):
                cell.value          = val
                cell.number_format  = "#,##0.00"
                cell.alignment      = Alignment(horizontal="right", vertical="center")
            elif col == "Discount %":
                cell.value          = val / 100
                cell.number_format  = "0%"
                cell.alignment      = Alignment(horizontal="center", vertical="center")
            elif col == "Units":
                cell.value          = int(val)
                cell.alignment      = Alignment(horizontal="right", vertical="center")
            else:
                cell.value          = str(val) if val is not None else ""
                cell.alignment      = Alignment(horizontal="left", vertical="center")
            cell.font   = Font(name="Arial", size=9)
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = _border()
        ws.row_dimensions[r].height = 15

    ws.freeze_panes = "B3"


# ── Main entry ────────────────────────────────────────────────────────────────
def load(data: dict, report_date: str = None) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    tag      = report_date or datetime.now().strftime("%Y_%m")
    out_path = OUTPUT_DIR / f"monthly_report_{tag}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Dashboard")
    ws2 = wb.create_sheet("Monthly Summary")
    ws3 = wb.create_sheet("Region Performance")
    ws4 = wb.create_sheet("Product Breakdown")
    ws5 = wb.create_sheet("Raw Data")

    _write_dashboard(ws1, data["kpis"], data["monthly"])
    _write_monthly(ws2, data["monthly"])
    _write_regions(ws3, data["region_perf"])
    _write_products(ws4, data["product_breakdown"], data["category_summary"])
    _write_raw(ws5, data["raw_completed"])

    ws1.sheet_properties.tabColor = DARK_BLUE
    ws2.sheet_properties.tabColor = MID_BLUE
    ws3.sheet_properties.tabColor = GREEN
    ws4.sheet_properties.tabColor = ACCENT
    ws5.sheet_properties.tabColor = "888888"

    wb.save(out_path)
    return out_path